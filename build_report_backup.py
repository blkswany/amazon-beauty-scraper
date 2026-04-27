#!/usr/bin/env python3
"""
Amazon Beauty Bestsellers Report Builder

기존 양식(260408_amazon_beauty.xlsx)과 동일한 구조로 새 리포트 생성
  1. 브랜드 시트  : 매핑 테이블 + 국가별 등장 횟수 수식(COUNTIF)
  2. 국가 시트    : rank/title/rating/reviews/기업명 (기업명은 Excel 수식)
  3. 한국 정리    : 한국 브랜드만 6개국 가로 배치

Usage:
    python build_report.py
    python build_report.py --scraped other_file.xlsx  # 다른 스크래핑 파일 지정
"""
import sys
import re
import argparse
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")

# ─── Config ───────────────────────────────────────────────────────────────────
TODAY = datetime.now().strftime("%Y-%m-%d")
TEMPLATE_FILE = Path("260408_amazon_beauty.xlsx")  # 브랜드 매핑 원본 (없어도 동작)

COUNTRIES = [
    ("US", "미국"),
    ("DE", "독일"),
    ("FR", "프랑스"),
    ("IT", "이탈리아"),
    ("UK", "영국"),
    ("ES", "스페인"),
]

# ─── 하드코딩 브랜드 매핑 (템플릿 파일 없을 때 기본값으로 사용) ──────────────
BUILTIN_BRAND_MAP = {
    "medicube":        "에이피알",
    "EQQUALBERRY":     "부스터스",
    "Dr.Althea":       "더퓨어랩",
    "Dr.Melaxin":      "브랜드501",
    "Anua":            "더파운더즈",
    "KAHI":            "코리아테크",
    "Biodance":        "뷰티셀렉션",
    "d'alba":          "달바글로벌",
    "celimax":         "앱솔브랩",
    "Cosrx":           "아모레퍼시픽",
    "Illiyoon":        "아모레퍼시픽",
    "MIZON":           "피에프디",
    "Beauty of Joseon":"구다이글로벌",
}

# ─── Style helpers ────────────────────────────────────────────────────────────
_THIN = Side(style="thin")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

FILL_TITLE    = PatternFill("solid", fgColor="1F4E79")
FILL_COUNTRY  = PatternFill("solid", fgColor="2E75B6")
FILL_HEADER   = PatternFill("solid", fgColor="BDD7EE")
FILL_KOREAN   = PatternFill("solid", fgColor="FFF2CC")
FILL_BRAND_H  = PatternFill("solid", fgColor="2E75B6")

def _cell(ws, row, col, value=None, fill=None, bold=False, font_color="000000",
          size=10, h_align="left", v_align="center", wrap=False, border=True, number_format=None):
    c = ws.cell(row=row, column=col)
    if value is not None:
        c.value = value
    if fill:
        c.fill = fill
    c.font = Font(bold=bold, color=font_color, size=size)
    c.alignment = Alignment(horizontal=h_align, vertical=v_align, wrap_text=wrap)
    if border:
        c.border = BORDER
    if number_format:
        c.number_format = number_format
    return c

# ─── Brand mapping ────────────────────────────────────────────────────────────
def load_brand_mapping():
    """
    브랜드 매핑 로드 순서:
    1. 템플릿 엑셀 파일(260408_amazon_beauty.xlsx)이 있으면 거기서 로드
    2. 없으면 하드코딩된 BUILTIN_BRAND_MAP 사용
    """
    if TEMPLATE_FILE.exists():
        try:
            df = pd.read_excel(TEMPLATE_FILE, sheet_name="브랜드", header=0)
            brands = {}
            for _, row in df.iterrows():
                brand = str(row.iloc[0]).strip()
                company = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ""
                if brand and company and brand not in ("브랜드명", "nan", "ㅇ"):
                    brands[brand] = company
            # 하드코딩 매핑도 병합 (템플릿에 없는 브랜드 보완)
            for brand, company in BUILTIN_BRAND_MAP.items():
                if brand not in brands:
                    brands[brand] = company
            print(f"  브랜드 매핑 {len(brands)}개 로드됨 (템플릿 파일 기준)")
            return brands
        except Exception as e:
            print(f"  [경고] 브랜드 시트 로드 실패: {e} — 기본 매핑 사용")

    print(f"  [안내] 템플릿 파일 없음 — 하드코딩 브랜드 매핑 {len(BUILTIN_BRAND_MAP)}개 사용")
    return dict(BUILTIN_BRAND_MAP)


def match_brand(title, brand_map):
    """타이틀에서 브랜드명 찾아 한국 기업명 반환 (없으면 빈 문자열)"""
    if not isinstance(title, str):
        return ""
    for brand, company in brand_map.items():
        if re.search(re.escape(brand), title, re.IGNORECASE):
            return company
    return ""

# ─── 브랜드 시트 ──────────────────────────────────────────────────────────────
def write_brand_sheet(wb, brand_map):
    ws = wb.create_sheet(title="브랜드")

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 18
    ws.row_dimensions[1].height = 22

    headers = ["브랜드명", "한국 기업명"]
    for col, h in enumerate(headers, 1):
        _cell(ws, 1, col, h, fill=FILL_BRAND_H, bold=True, font_color="FFFFFF",
              h_align="center")

    for r, (brand, company) in enumerate(brand_map.items(), 2):
        ws.row_dimensions[r].height = 16
        _cell(ws, r, 1, brand, h_align="left")
        _cell(ws, r, 2, company, bold=True, h_align="center")

    print(f"  [브랜드] 시트 완료 ({len(brand_map)}개 브랜드)")
    return ws

# ─── 국가 시트 ────────────────────────────────────────────────────────────────
def write_country_sheet(wb, country_en, country_ko, df_raw, brand_map):
    ws = wb.create_sheet(title=country_ko)

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 90
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14
    ws.row_dimensions[1].height = 20

    hdrs = ["rank", "title", "rating", "reviews", "기업명"]
    for col, h in enumerate(hdrs, 1):
        _cell(ws, 1, col, h, fill=FILL_BRAND_H, bold=True,
              font_color="FFFFFF", h_align="center")

    for i, row in df_raw.iterrows():
        r = i + 2
        title   = row.get("title", "")
        rating  = row.get("rating", "")
        reviews = row.get("reviews", "")
        company = match_brand(title, brand_map)

        ws.row_dimensions[r].height = 14
        fill = FILL_KOREAN if company else None

        _cell(ws, r, 1, row["rank"], fill=fill, h_align="center")
        _cell(ws, r, 2, title,       fill=fill, wrap=True)
        _cell(ws, r, 3, rating,      fill=fill, h_align="center")
        _cell(ws, r, 4, reviews,     fill=fill, h_align="center")
        _cell(ws, r, 5, company,     fill=fill, h_align="center", bold=bool(company))

    ws.freeze_panes = "A2"
    print(f"  [{country_ko}] 시트 완료 ({len(df_raw)}행)")
    return ws

# ─── 한국 정리 색상 팔레트 ────────────────────────────────────────────────────
GRAY_TITLE   = PatternFill("solid", fgColor="404040")
GRAY_COUNTRY = PatternFill("solid", fgColor="737373")
GRAY_HEADER  = PatternFill("solid", fgColor="D9D9D9")
GRAY_ROW_ODD = PatternFill("solid", fgColor="FFFFFF")
GRAY_ROW_EVN = PatternFill("solid", fgColor="F2F2F2")
GRAY_COMPANY = PatternFill("solid", fgColor="E8E8E8")
GRAY_BORDER  = Side(style="thin", color="BFBFBF")

def _gray_border():
    return Border(left=GRAY_BORDER, right=GRAY_BORDER,
                  top=GRAY_BORDER, bottom=GRAY_BORDER)

def _kcell(ws, row, col, value=None, fill=None, bold=False,
           font_color="2D2D2D", size=9, h_align="left", wrap=False):
    c = ws.cell(row=row, column=col)
    if value is not None:
        c.value = value
    if fill:
        c.fill = fill
    c.font = Font(bold=bold, color=font_color, size=size)
    c.alignment = Alignment(horizontal=h_align, vertical="center",
                             wrap_text=wrap, shrink_to_fit=False)
    c.border = _gray_border()
    return c

# ─── 한국 정리 시트 ───────────────────────────────────────────────────────────
def write_korean_summary(wb, country_data, brand_map):
    ws = wb.create_sheet(title="한국 정리")
    COLS_PER = 5
    COL_GAP  = 1
    STRIDE   = COLS_PER + COL_GAP

    korean_by_country = []
    for en, ko in COUNTRIES:
        df = country_data.get(en, pd.DataFrame())
        rows = []
        for _, r in df.iterrows():
            company = match_brand(r.get("title", ""), brand_map)
            if company:
                rows.append({
                    "rank":    r["rank"],
                    "title":   r.get("title", ""),
                    "rating":  r.get("rating", ""),
                    "reviews": r.get("reviews", ""),
                    "company": company,
                })
        korean_by_country.append(rows)

    max_rows = max((len(r) for r in korean_by_country), default=0)

    for ci in range(len(COUNTRIES)):
        base = ci * STRIDE + 1
        ws.column_dimensions[get_column_letter(base)].width   = 5
        ws.column_dimensions[get_column_letter(base+1)].width = 48
        ws.column_dimensions[get_column_letter(base+2)].width = 6
        ws.column_dimensions[get_column_letter(base+3)].width = 9
        ws.column_dimensions[get_column_letter(base+4)].width = 13
        if ci < len(COUNTRIES) - 1:
            ws.column_dimensions[get_column_letter(base+5)].width = 2

    ws.row_dimensions[1].height = 20
    for ci, (en, ko) in enumerate(COUNTRIES):
        base = ci * STRIDE + 1
        ws.merge_cells(start_row=1, start_column=base,
                       end_row=1,   end_column=base + COLS_PER - 1)
        _kcell(ws, 1, base, "Amazon Best Sellers in Beauty",
               fill=GRAY_TITLE, bold=True, font_color="FFFFFF",
               h_align="center", size=9)

    ws.row_dimensions[2].height = 18
    for ci, (en, ko) in enumerate(COUNTRIES):
        base = ci * STRIDE + 1
        ws.merge_cells(start_row=2, start_column=base,
                       end_row=2,   end_column=base + COLS_PER - 1)
        _kcell(ws, 2, base, ko,
               fill=GRAY_COUNTRY, bold=True, font_color="FFFFFF",
               h_align="center", size=10)

    ws.row_dimensions[3].height = 16
    col_labels = ["순위", "제품", "평점", "리뷰", "회사"]
    for ci in range(len(COUNTRIES)):
        base = ci * STRIDE + 1
        for j, lbl in enumerate(col_labels):
            _kcell(ws, 3, base + j, lbl,
                   fill=GRAY_HEADER, bold=True, h_align="center", size=9)

    for idx in range(max_rows):
        r = idx + 4
        ws.row_dimensions[r].height = 16
        row_fill = GRAY_ROW_ODD if idx % 2 == 0 else GRAY_ROW_EVN

        for ci, rows in enumerate(korean_by_country):
            base = ci * STRIDE + 1
            if idx < len(rows):
                item = rows[idx]
                _kcell(ws, r, base,   item["rank"],    fill=row_fill, h_align="center", bold=True)
                _kcell(ws, r, base+1, item["title"],   fill=row_fill, h_align="left", wrap=False)
                _kcell(ws, r, base+2, item["rating"],  fill=row_fill, h_align="center")
                _kcell(ws, r, base+3, item["reviews"], fill=row_fill, h_align="center")
                _kcell(ws, r, base+4, item["company"], fill=GRAY_COMPANY, h_align="center", bold=True)
            else:
                for j in range(COLS_PER):
                    _kcell(ws, r, base + j, fill=row_fill)

    ws.freeze_panes = "A4"

    counts = [len(r) for r in korean_by_country]
    print(f"  [한국 정리] 시트 완료 — " +
          ", ".join(f"{ko}:{n}개" for (_, ko), n in zip(COUNTRIES, counts)))

# ─── Main ─────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--scraped", default=f"amazon_beauty_bestsellers_{TODAY}.xlsx",
                        help="스크래핑된 xlsx 파일 경로")
    parser.add_argument("--output", default=f"amazon_beauty_report_{TODAY}.xlsx",
                        help="출력 파일 경로")
    args = parser.parse_args()

    scraped_path = Path(args.scraped)
    output_path  = Path(args.output)

    print("=" * 55)
    print(" Amazon Beauty Report Builder")
    print(f" 입력: {scraped_path}")
    print(f" 출력: {output_path}")
    print("=" * 55)

    if not scraped_path.exists():
        print(f"[오류] 파일 없음: {scraped_path}")
        sys.exit(1)

    country_data = {}
    for en, ko in COUNTRIES:
        try:
            df = pd.read_excel(scraped_path, sheet_name=en)
            country_data[en] = df
            print(f"  [{en}] {len(df)}행 로드")
        except Exception as e:
            print(f"  [{en}] 로드 실패: {e}")
            country_data[en] = pd.DataFrame()

    brand_map = load_brand_mapping()

    wb = Workbook()
    wb.remove(wb.active)

    write_brand_sheet(wb, brand_map)

    for en, ko in COUNTRIES:
        df = country_data.get(en, pd.DataFrame())
        if df.empty:
            wb.create_sheet(title=ko)
            print(f"  [{ko}] 빈 시트 생성 (데이터 없음)")
        else:
            write_country_sheet(wb, en, ko, df, brand_map)

    write_korean_summary(wb, country_data, brand_map)

    try:
        wb.save(output_path)
    except PermissionError:
        ts = datetime.now().strftime("%H%M%S")
        output_path = output_path.with_stem(output_path.stem + f"_{ts}")
        wb.save(output_path)
        print(f"  [주의] 기존 파일이 열려 있어 새 이름으로 저장됨")
    print(f"\n완료! → {output_path.resolve()}")


if __name__ == "__main__":
    main()
