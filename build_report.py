#!/usr/bin/env python3
"""
Amazon & Rakuten Beauty Report Builder
영어 및 일본어 브랜드 매핑 지원 버전
"""
import sys
import re
import argparse
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")

# ─── Config ───────────────────────────────────────────────────────────────────
TODAY = datetime.now().strftime("%Y-%m-%d")
TEMPLATE_FILE = Path("260408_amazon_beauty.xlsx")  

# 리포트에 포함될 시트 구성
COUNTRIES = [
    ("US", "미국(전체)"),
    ("US_SkinCare", "미국(스킨케어)"),
    ("US_SunCare", "미국(선크림)"),
    ("Rakuten_JP", "일본(라쿠텐)"),
    ("Qoo10_JP", "일본(Qoo10)"),
    ("DE", "독일"),
    ("FR", "프랑스"),
    ("IT", "이탈리아"),
    ("UK", "영국"),
    ("ES", "스페인"),
]

# ─── 브랜드 매핑 (영어 & 일본어 통합) ──────────────────────────────────────────
# 라쿠텐 상품명에서 자주 보이는 일본어 표기를 추가했습니다.
BUILTIN_BRAND_MAP = {
    # ─── 영어 키 (Amazon / 영문 타이틀 매칭용) ─────────────────────
    "medicube":         "에이피알",
    "APRILSKIN":        "에이피알",
    "April Skin":       "에이피알",
    "EQQUALBERRY":      "부스터스",
    "Dr.Althea":        "더퓨어랩",
    "Dr.Melaxin":       "브랜드501",
    "Anua":             "더파운더즈",
    "KAHI":             "코리아테크",
    "Biodance":         "뷰티셀렉션",
    "d'alba":           "달바글로벌",
    "d'Alba":           "달바글로벌",
    "celimax":          "앱솔브랩",
    "Cosrx":            "아모레퍼시픽",
    "Illiyoon":         "아모레퍼시픽",
    "AESTURA":          "아모레퍼시픽",
    "ESTRA":            "아모레퍼시픽",
    "MIZON":            "피에프디",
    "Beauty of Joseon": "구다이글로벌",
    "TIRTIR":           "구다이글로벌",
    "VT":               "브이티",
    "manyo":            "마녀공장",
    "Manyo":            "마녀공장",
    "paparecipe":       "코스토리",
    "ZEROID":           "네오팜",

    # ─── 일본어 카타카나 / 한자 표기 (Rakuten·Qoo10 일본 시트) ─────
    # 더파운더즈
    "アヌア":           "더파운더즈",
    # 달바글로벌
    "ダルバ":           "달바글로벌",
    "ホワイトトリュフファーストスプレー": "달바글로벌",   # d'Alba 시그니처 제품명
    # 티르티르
    "ティルティル":      "구다이글로벌",
    "マスクフィット":    "구다이글로벌",                       # TIRTIR MaskFit 라인
    # 마녀공장
    "魔女工場":         "마녀공장",
    "マニョ":           "마녀공장",
    # 에이피알
    "メディキューブ":    "에이피알",
    "エイプリルスキン":  "에이피알",
    # 코리아테크
    "カヒ":             "코리아테크",
    # 뷰티셀렉션
    "バイオダンス":      "뷰티셀렉션",
    # 구다이글로벌
    "朝鮮美女":         "구다이글로벌",
    # 아모레퍼시픽 산하
    "コスアールエックス": "아모레퍼시픽",                  # COSRX
    "イニスフリー":      "아모레퍼시픽",                  # innisfree
    "ラネージュ":        "아모레퍼시픽",                  # LANEIGE
    "雪花秀":           "아모레퍼시픽",                  # Sulwhasoo
    "アトバリア":        "아모레퍼시픽",                  # Atobarrier (ESTRA)
    "エストラ":          "아모레퍼시픽",                  # ESTRA
    # 파파레시피
    "パパレシピ":        "코스토리",
    # 네오팜
    "リアルバリア":      "네오팜",                       # Real Barrier
    "エクストリームクリームマスク": "네오팜",              # Real Barrier 시그니처
    # 동국제약 (센텔리안24)
    "マデカクリーム":    "동국제약",                     # Madeca (Centellian 24)
    # 주의: "マデカ" 단독은 사용 금지 — "マデカッソシド"(성분명)에 오매칭됨
    # 비나우
    "ナンバーズイン":    "비나우",                       # numbuzin
    # 토리든
    "トリデン":          "토리든",
    # 아이패밀리에스씨
    "ロムアンド":        "아이패밀리에스씨",              # rom&nd
    # 클리오
    "クリオ":            "클리오",
    "ペリペラ":          "클리오",                       # peripera
}

# ─── Style & Helpers ──────────────────────────────────────────────────────────
_THIN = Side(style="thin")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
FILL_BRAND_H  = PatternFill("solid", fgColor="2E75B6")
FILL_KOREAN   = PatternFill("solid", fgColor="FFF2CC")

def _cell(ws, row, col, value=None, fill=None, bold=False, font_color="000000",
          size=10, h_align="left", v_align="center", wrap=False, border=True):
    c = ws.cell(row=row, column=col)
    if value is not None: c.value = value
    if fill: c.fill = fill
    c.font = Font(bold=bold, color=font_color, size=size)
    c.alignment = Alignment(horizontal=h_align, vertical=v_align, wrap_text=wrap)
    if border: c.border = BORDER
    return c

def load_brand_mapping():
    """템플릿 파일이 있으면 로드하고, 없으면 내장 맵을 반환"""
    brands = dict(BUILTIN_BRAND_MAP)
    if TEMPLATE_FILE.exists():
        try:
            df = pd.read_excel(TEMPLATE_FILE, sheet_name="브랜드", header=0)
            for _, row in df.iterrows():
                brand = str(row.iloc[0]).strip()
                company = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ""
                if brand and company and brand not in ("브랜드명", "nan"):
                    brands[brand] = company
            print(f"  브랜드 매핑 로드 완료 (템플릿 + 내장 목록)")
        except Exception as e:
            print(f"  [경고] 템플릿 로드 실패: {e}")
    return brands

def match_brand(title, brand_map, brand=""):
    """브랜드 필드 직접 매칭 → 타이틀 서브스트링 매칭 순으로 한국 기업명 반환"""
    if brand and isinstance(brand, str):
        for key, company in brand_map.items():
            if key.lower() == brand.lower():
                return company
    if not isinstance(title, str):
        return ""
    for key, company in brand_map.items():
        if re.search(re.escape(key), title, re.IGNORECASE):
            return company
    return ""

# ─── 시트 작성 함수들 ─────────────────────────────────────────────────────────
def write_brand_sheet(wb, brand_map):
    ws = wb.create_sheet(title="브랜드")
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 18
    _cell(ws, 1, 1, "브랜드명(Key)", fill=FILL_BRAND_H, bold=True, font_color="FFFFFF", h_align="center")
    _cell(ws, 1, 2, "한국 기업명", fill=FILL_BRAND_H, bold=True, font_color="FFFFFF", h_align="center")
    for r, (brand, company) in enumerate(brand_map.items(), 2):
        _cell(ws, r, 1, brand)
        _cell(ws, r, 2, company, bold=True, h_align="center")

def write_country_sheet(wb, country_ko, df_raw, brand_map, country_en=""):
    ws = wb.create_sheet(title=country_ko)
    is_qoo10 = country_en == "Qoo10_JP"

    if is_qoo10:
        widths = [6, 90, 12, 14]
        hdrs = ["rank", "title", "reviews", "기업명"]
    else:
        widths = [6, 90, 8, 12, 14]
        hdrs = ["rank", "title", "rating", "reviews", "기업명"]

    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    for col, h in enumerate(hdrs, 1):
        _cell(ws, 1, col, h, fill=FILL_BRAND_H, bold=True, font_color="FFFFFF", h_align="center")

    for i, row in df_raw.iterrows():
        r = i + 2
        title = row.get("title", "")
        company = match_brand(title, brand_map, brand=row.get("brand", ""))
        fill = FILL_KOREAN if company else None

        _cell(ws, r, 1, row.get("rank", ""), fill=fill, h_align="center")
        _cell(ws, r, 2, title, fill=fill, wrap=True)
        if is_qoo10:
            _cell(ws, r, 3, row.get("reviews", ""), fill=fill, h_align="center")
            _cell(ws, r, 4, company, fill=fill, h_align="center", bold=bool(company))
        else:
            _cell(ws, r, 3, row.get("rating", ""), fill=fill, h_align="center")
            _cell(ws, r, 4, row.get("reviews", ""), fill=fill, h_align="center")
            _cell(ws, r, 5, company, fill=fill, h_align="center", bold=bool(company))
    ws.freeze_panes = "A2"

# ─── 한국 정리 시트 (디자인은 이전과 동일) ──────────────────────────────────────
GRAY_TITLE   = PatternFill("solid", fgColor="404040")
GRAY_COUNTRY = PatternFill("solid", fgColor="737373")
GRAY_HEADER  = PatternFill("solid", fgColor="D9D9D9")
GRAY_COMPANY = PatternFill("solid", fgColor="E8E8E8")
GRAY_BORDER  = Side(style="thin", color="BFBFBF")

def _kcell(ws, row, col, value=None, fill=None, bold=False, font_color="2D2D2D", size=9, h_align="left", wrap=False):
    c = ws.cell(row=row, column=col)
    if value is not None: c.value = value
    if fill: c.fill = fill
    c.font = Font(bold=bold, color=font_color, size=size)
    c.alignment = Alignment(horizontal=h_align, vertical="center", wrap_text=wrap)
    c.border = Border(left=GRAY_BORDER, right=GRAY_BORDER, top=GRAY_BORDER, bottom=GRAY_BORDER)
    return c

def write_korean_summary(wb, country_data, brand_map):
    ws = wb.create_sheet(title="한국 정리")

    # 국가별 컬럼 구성: Qoo10은 평점 제외 (4컬럼), 그 외는 5컬럼
    def cols_of(en):
        if en == "Qoo10_JP":
            return {"labels": ["순위", "제품", "리뷰", "회사"],
                    "widths": [5, 48, 9, 13]}
        return {"labels": ["순위", "제품", "평점", "리뷰", "회사"],
                "widths": [5, 48, 6, 9, 13]}

    korean_by_country = []
    for en, ko in COUNTRIES:
        df = country_data.get(en, pd.DataFrame())
        rows = [{"rank": r.get("rank",""), "title": r.get("title",""), "rating": r.get("rating",""),
                 "reviews": r.get("reviews",""), "company": match_brand(r.get("title",""), brand_map, brand=r.get("brand",""))}
                for _, r in df.iterrows() if match_brand(r.get("title",""), brand_map, brand=r.get("brand",""))]
        korean_by_country.append(rows)

    # 국가별 base 컬럼 위치 누적 계산 (컬럼 수 + 1칸 갭)
    bases = []
    offset = 1
    for en, ko in COUNTRIES:
        bases.append(offset)
        offset += len(cols_of(en)["labels"]) + 1

    max_rows = max((len(r) for r in korean_by_country), default=0)
    for ci, (en, ko) in enumerate(COUNTRIES):
        base = bases[ci]
        cfg = cols_of(en)
        n = len(cfg["labels"])
        for j, w in enumerate(cfg["widths"]):
            ws.column_dimensions[get_column_letter(base+j)].width = w
        ws.merge_cells(start_row=1, start_column=base, end_row=1, end_column=base+n-1)
        _kcell(ws, 1, base, "Best Sellers in Beauty", fill=GRAY_TITLE, bold=True, font_color="FFFFFF", h_align="center")
        ws.merge_cells(start_row=2, start_column=base, end_row=2, end_column=base+n-1)
        _kcell(ws, 2, base, ko, fill=GRAY_COUNTRY, bold=True, font_color="FFFFFF", h_align="center")
        for j, lbl in enumerate(cfg["labels"]):
            _kcell(ws, 3, base+j, lbl, fill=GRAY_HEADER, bold=True, h_align="center")

    for idx in range(max_rows):
        r = idx + 4
        row_fill = PatternFill("solid", fgColor="FFFFFF") if idx % 2 == 0 else PatternFill("solid", fgColor="F2F2F2")
        for ci, (en, ko) in enumerate(COUNTRIES):
            base = bases[ci]
            rows = korean_by_country[ci]
            n = len(cols_of(en)["labels"])
            if idx < len(rows):
                item = rows[idx]
                _kcell(ws, r, base, item["rank"], fill=row_fill, h_align="center", bold=True)
                _kcell(ws, r, base+1, item["title"], fill=row_fill)
                if en == "Qoo10_JP":
                    _kcell(ws, r, base+2, item["reviews"], fill=row_fill, h_align="center")
                    _kcell(ws, r, base+3, item["company"], fill=GRAY_COMPANY, h_align="center", bold=True)
                else:
                    _kcell(ws, r, base+2, item["rating"], fill=row_fill, h_align="center")
                    _kcell(ws, r, base+3, item["reviews"], fill=row_fill, h_align="center")
                    _kcell(ws, r, base+4, item["company"], fill=GRAY_COMPANY, h_align="center", bold=True)
            else:
                for j in range(n): _kcell(ws, r, base+j, fill=row_fill)
    ws.freeze_panes = "A4"

# ─── Main ─────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--scraped", default=f"amazon_beauty_bestsellers_{TODAY}.xlsx")
    parser.add_argument("--output", default=f"amazon_beauty_report_{TODAY}.xlsx")
    args = parser.parse_args()

    if not Path(args.scraped).exists():
        print(f"[오류] 파일 없음: {args.scraped}")
        sys.exit(1)

    brand_map = load_brand_mapping()
    country_data = {en: pd.read_excel(args.scraped, sheet_name=en) for en, ko in COUNTRIES}
    
    wb = Workbook()
    wb.remove(wb.active)
    write_brand_sheet(wb, brand_map)
    for en, ko in COUNTRIES:
        write_country_sheet(wb, ko, country_data[en], brand_map, country_en=en)
    write_korean_summary(wb, country_data, brand_map)
    
    wb.save(args.output)
    print(f"\n[성공] 리포트 생성 완료: {args.output}")

if __name__ == "__main__":
    main()
